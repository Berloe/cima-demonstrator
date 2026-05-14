"""FileProcessingAdapter → FileProcessingPort (KIMA_Infrastructure_Layer_v0.6 §3.8)."""
from __future__ import annotations

import csv
import io
import json
import logging
import re

from cima_demo.domain.errors import FileProcessingError
from cima_demo.domain.ports import FileProcessingPort

log = logging.getLogger(__name__)

# Extensions treated as source code
_CODE_EXTS = frozenset({
    "py", "js", "ts", "jsx", "tsx", "java", "c", "cpp", "h", "hpp",
    "go", "rs", "rb", "php", "sh", "bash", "zsh", "sql", "yaml", "yml",
    "toml", "ini", "cfg", "xml", "html", "htm", "css", "scss",
    "kt", "swift", "cs", "dart", "r", "lua", "pl", "vim",
})

_SUPPORTED_MIME = [
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/json",
    "application/zip",
    "application/x-zip-compressed",
    "text/plain",
    "text/markdown",
    "text/csv",
    "text/html",
    "image/png",
    "image/jpeg",
    "image/tiff",
    "image/webp",
]


def infer_doc_type(filename: str, mime_type: str) -> str:
    """Return a short normalized doc type from filename extension and mime type.

    Returns one of: "pdf" | "docx" | "markdown" | "csv" | "json" | "html" |
                    "code" | "text" | "image"
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "zip" or mime_type in (
        "application/zip", "application/x-zip-compressed", "application/x-zip",
    ):
        return "zip"
    if mime_type == "application/pdf" or ext == "pdf":
        return "pdf"
    if (mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or ext == "docx"):
        return "docx"
    if (mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ) or ext in ("xlsx", "xls")):
        return "xlsx"
    if ext in ("md", "markdown") or mime_type == "text/markdown":
        return "markdown"
    if ext == "csv" or mime_type == "text/csv":
        return "csv"
    if ext == "json" or mime_type == "application/json":
        return "json"
    if ext in ("html", "htm") or mime_type in ("text/html", "application/xhtml+xml"):
        return "html"
    if mime_type.startswith("image/") or ext in ("png", "jpg", "jpeg", "tiff", "tif", "webp"):
        return "image"
    if ext in _CODE_EXTS:
        return "code"
    return "text"


class FileProcessingAdapter(FileProcessingPort):
    """Extracts text from PDF, DOCX, plain text, CSV, JSON, and code files.

    PDF:  embeds [PAGE N] markers so the chunker can track page numbers.
    DOCX: extracts paragraphs AND table cells with structural markers.
    CSV:  converts rows to readable "key: value" blocks (max 500 rows).
    JSON: pretty-prints with 8 000-char cap.
    Code: adds a language header line.
    HTML: strips tags to plain text.
    """

    def extract_text(
        self,
        content: bytes,
        filename: str,
        mime_type: str,
    ) -> str:
        try:
            doc_type = infer_doc_type(filename, mime_type)
            if doc_type == "pdf":
                return self._extract_pdf(content)
            if doc_type == "docx":
                return self._extract_docx(content)
            if doc_type == "xlsx":
                return self._extract_xlsx(content, filename)
            if doc_type == "csv":
                return self._extract_csv(content, filename)
            if doc_type == "json":
                return self._extract_json(content, filename)
            if doc_type == "html":
                return self._extract_html(content)
            if doc_type == "image":
                return self._extract_image_ocr(content, filename)
            if doc_type == "code":
                ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
                return f"# [{filename}] ({ext})\n\n" + content.decode("utf-8", errors="replace")
            # markdown / text / fallback
            return content.decode("utf-8", errors="replace")
        except FileProcessingError:
            raise
        except Exception as exc:
            raise FileProcessingError(f"Failed to extract text from {filename}: {exc}") from exc

    # ── PDF ───────────────────────────────────────────────────────────────────

    def _extract_pdf(self, content: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise FileProcessingError("pypdf not installed") from exc

        reader = PdfReader(io.BytesIO(content))
        parts: list[str] = []
        for page_num, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                parts.append(f"[PAGE {page_num}]\n{text.strip()}")

        result = "\n\n".join(parts)

        # Scanned PDF: no selectable text — fall back to per-page OCR (FFT-01)
        if not result.strip():
            result = self._ocr_pdf_pages(reader)

        return result

    def _ocr_pdf_pages(self, reader: object) -> str:
        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            log.warning("OCR fallback unavailable for scanned PDF: install pytesseract + Pillow.")
            return ""
        parts: list[str] = []
        for page_num, page in enumerate(getattr(reader, "pages", []), start=1):
            page_parts: list[str] = []
            for image_file in page.images:
                try:
                    img = Image.open(io.BytesIO(image_file.data))
                    page_parts.append(str(pytesseract.image_to_string(img)))
                except Exception as exc:
                    log.debug("OCR failed for PDF page image: %s", exc)
            if page_parts:
                parts.append(f"[PAGE {page_num}]\n" + "\n".join(p for p in page_parts if p.strip()))
        return "\n\n".join(p for p in parts if p.strip())

    # ── DOCX ─────────────────────────────────────────────────────────────────

    def _extract_docx(self, content: bytes) -> str:
        try:
            from docx import Document
            from docx.table import Table
        except ImportError as exc:
            raise FileProcessingError("python-docx not installed") from exc

        doc = Document(io.BytesIO(content))
        parts: list[str] = []

        # Iterate document body children in order (paragraphs and tables)
        for block in doc.element.body:
            tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
            if tag == "p":
                # Paragraph
                from docx.text.paragraph import Paragraph
                para = Paragraph(block, doc)
                text = para.text.strip()
                if text:
                    parts.append(text)
            elif tag == "tbl":
                # Table — render as simple grid text
                from docx.table import Table as DocxTable
                table = DocxTable(block, doc)
                table_lines: list[str] = []
                for row in table.rows:
                    cells = [cell.text.strip() for cell in row.cells]
                    # Deduplicate merged cells (python-docx repeats merged cells)
                    seen: list[str] = []
                    for c in cells:
                        if not seen or seen[-1] != c:
                            seen.append(c)
                    table_lines.append(" | ".join(seen))
                if table_lines:
                    parts.append("[TABLE]\n" + "\n".join(table_lines))

        return "\n\n".join(parts)

    # ── XLSX ─────────────────────────────────────────────────────────────────

    def _extract_xlsx(self, content: bytes, filename: str) -> str:
        """Convert Excel workbook to readable text.

        Each sheet is rendered as a tab-separated table with a [SHEET: name] header.
        Max 1000 rows per sheet to avoid context overflow.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise FileProcessingError("openpyxl not installed") from exc

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_written = 0
            sheet_lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                if rows_written >= 1000:
                    sheet_lines.append(f"[...truncated — sheet has more than 1000 rows]")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                # Skip entirely empty rows
                if any(c.strip() for c in cells):
                    sheet_lines.append("\t".join(cells))
                    rows_written += 1
            if sheet_lines:
                parts.append(f"[SHEET: {sheet_name}]\n" + "\n".join(sheet_lines))
        wb.close()

        header = f"[XLSX: {filename}] {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}\n\n"
        return header + "\n\n".join(parts)

    # ── CSV ──────────────────────────────────────────────────────────────────

    def _extract_csv(self, content: bytes, filename: str) -> str:
        """Convert CSV to readable key:value blocks (max 500 rows)."""
        text = content.decode("utf-8", errors="replace")
        lines: list[str] = []
        try:
            reader = csv.DictReader(io.StringIO(text))
            fields = reader.fieldnames or []
            rows_written = 0
            for row in reader:
                if rows_written >= 500:
                    lines.append(f"[...truncated — file has more than 500 rows]")
                    break
                block = "\n".join(f"{k}: {v}" for k, v in row.items() if v is not None)
                lines.append(block)
                rows_written += 1
            header = f"[CSV: {filename}] {len(fields)} columns: {', '.join(str(f) for f in fields)}"
            return header + "\n\n" + "\n\n".join(lines)
        except Exception:
            # Fallback: return raw text
            return text

    # ── JSON ─────────────────────────────────────────────────────────────────

    def _extract_json(self, content: bytes, filename: str) -> str:
        text = content.decode("utf-8", errors="replace")
        try:
            parsed = json.loads(text)
            pretty = json.dumps(parsed, indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            pretty = text
        cap = 8000
        header = f"[JSON: {filename}]\n\n"
        if len(pretty) > cap:
            return header + pretty[:cap] + f"\n\n[...truncated — {len(pretty) - cap} more chars]"
        return header + pretty

    # ── HTML ─────────────────────────────────────────────────────────────────

    def _extract_html(self, content: bytes) -> str:
        """Strip HTML tags to plain text preserving paragraph structure.

        Block-level elements (p, div, h1-h6, li, td…) emit a paragraph break
        so the chunker can segment the document semantically.
        """
        from html.parser import HTMLParser

        class _Stripper(HTMLParser):
            _SKIP = frozenset({"script", "style", "noscript", "template"})
            _BLOCK = frozenset({
                "p", "div", "article", "section", "main", "aside",
                "h1", "h2", "h3", "h4", "h5", "h6",
                "li", "td", "th", "dt", "dd",
                "blockquote", "pre", "br", "hr",
                "header", "footer", "nav",
            })

            def __init__(self) -> None:
                super().__init__()
                self._paragraphs: list[str] = []
                self._block: list[str] = []
                self._skip_depth = 0

            def _flush_block(self) -> None:
                text = " ".join(self._block).strip()
                if text:
                    self._paragraphs.append(text)
                self._block = []

            def handle_starttag(self, tag: str, attrs: list) -> None:  # type: ignore[override]
                if tag in self._SKIP:
                    self._skip_depth += 1
                elif tag in self._BLOCK and self._skip_depth == 0:
                    self._flush_block()

            def handle_endtag(self, tag: str) -> None:
                if tag in self._SKIP:
                    self._skip_depth = max(0, self._skip_depth - 1)
                elif tag in self._BLOCK and self._skip_depth == 0:
                    self._flush_block()

            def handle_data(self, data: str) -> None:
                if self._skip_depth == 0 and data.strip():
                    self._block.append(data.strip())

            @property
            def text(self) -> str:
                self._flush_block()
                return "\n\n".join(self._paragraphs)

        text = content.decode("utf-8", errors="replace")
        stripper = _Stripper()
        stripper.feed(text)
        return stripper.text

    # ── Image OCR ─────────────────────────────────────────────────────────────

    def _extract_image_ocr(self, content: bytes, filename: str) -> str:
        """OCR via pytesseract + Pillow (FFT-01). Graceful degradation if not installed."""
        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            log.warning(
                "OCR unavailable for %s: install pytesseract + Pillow "
                "(pip install pytesseract Pillow) and Tesseract-OCR.",
                filename,
            )
            return ""
        image = Image.open(io.BytesIO(content))
        return str(pytesseract.image_to_string(image))

    def supported_mime_types(self) -> list[str]:
        return list(_SUPPORTED_MIME)
